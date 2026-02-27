import csv
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date
from typing import List, Optional, Dict, Any
from dataclasses import dataclass
from enum import Enum
import os
import sys

# ---------- Константы и перечисления ----------
class UserRole(str, Enum):
    MANAGER = "Менеджер"
    MECHANIC = "Автомеханик"
    OPERATOR = "Оператор"
    CUSTOMER = "Заказчик"

class RequestStatus(str, Enum):
    NEW = "Новая заявка"
    IN_PROGRESS = "В процессе ремонта"
    WAITING_PARTS = "Ожидание автозапчастей"
    READY = "Готова к выдаче"

# ---------- Классы данных ----------
@dataclass
class User:
    user_id: int
    fio: str
    phone: str
    login: str
    password: str
    role: UserRole

@dataclass
class Request:
    request_id: int
    start_date: date
    car_type: str
    car_model: str
    problem_desc: str
    status: RequestStatus
    completion_date: Optional[date]
    repair_parts: Optional[str]
    master_id: Optional[int]
    client_id: int

@dataclass
class Comment:
    comment_id: int
    message: str
    master_id: int
    request_id: int

# ---------- Функции для диалогового выбора файла ----------
def ask_file(prompt: str, filetypes: list) -> Optional[str]:
    root = tk.Tk()
    root.withdraw()
    filename = filedialog.askopenfilename(title=prompt, filetypes=filetypes)
    root.destroy()
    return filename if filename else None

# ---------- Вспомогательные функции для чтения файлов ----------
def find_data_file(basename: str) -> Optional[str]:
    """Ищет файл с именем basename и расширением .csv, .txt или .xlsx (без учёта регистра)."""
    for ext in ['.csv', '.txt', '.xlsx']:
        target = basename + ext
        for filename in os.listdir('.'):
            if filename.lower() == target.lower():
                return filename
    return None

def check_openpyxl():
    try:
        import openpyxl
        return True
    except ImportError:
        return False

def read_excel_file(filename: str) -> Optional[List[Dict[str, Any]]]:
    if not check_openpyxl():
        messagebox.showerror("Ошибка", f"Для чтения Excel требуется openpyxl. Установите: pip install openpyxl")
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename, data_only=True)
        sheet = wb.active
        if sheet is None:
            return []
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        val = row[i]
                        if isinstance(val, str):
                            val = val.strip()
                        row_dict[header] = val
                    else:
                        row_dict[header] = None
                rows.append(row_dict)
        return rows
    except Exception as e:
        messagebox.showerror("Ошибка Excel", f"Не удалось прочитать файл {filename}:\n{e}")
        return None

def read_csv_file(filename: str) -> Optional[List[Dict[str, str]]]:
    encodings = ['utf-8-sig', 'utf-8', 'cp1251']
    for enc in encodings:
        try:
            with open(filename, 'r', encoding=enc) as f:
                # Определяем разделитель (по умолчанию ';', но если там ',', подстроимся)
                sample = f.read(1024)
                f.seek(0)
                if ';' in sample:
                    delimiter = ';'
                elif ',' in sample:
                    delimiter = ','
                else:
                    delimiter = ';'  # на всякий случай

                reader = csv.DictReader(f, delimiter=delimiter, skipinitialspace=True)
                # Приводим заголовки к нижнему регистру и убираем пробелы
                reader.fieldnames = [h.strip().lower() for h in reader.fieldnames if h]
                rows = []
                for row in reader:
                    cleaned = {k: v.strip() if isinstance(v, str) else v for k, v in row.items()}
                    rows.append(cleaned)
                if rows:
                    return rows
                return []
        except (UnicodeDecodeError, csv.Error):
            continue
    messagebox.showerror("Ошибка", f"Не удалось прочитать файл {filename} ни в одной кодировке.")
    return None

def load_data_file_interactive(basename: str, description: str) -> Optional[List[Dict[str, Any]]]:
    filename = find_data_file(basename)
    if filename:
        ext = os.path.splitext(filename)[1].lower()
        if ext == '.xlsx':
            return read_excel_file(filename)
        else:
            return read_csv_file(filename)

    filetypes = [
        ("Все поддерживаемые", "*.csv;*.txt;*.xlsx"),
        ("CSV файлы", "*.csv"),
        ("Текстовые файлы", "*.txt"),
        ("Excel файлы", "*.xlsx")
    ]
    msg = (f"Файл {basename} не найден в текущей папке.\n"
           f"Пожалуйста, выберите файл с данными {description} вручную.")
    messagebox.showinfo("Выбор файла", msg)

    filename = ask_file(f"Выберите файл {basename}", filetypes)
    if not filename:
        messagebox.showerror("Ошибка", "Файл не выбран. Программа будет закрыта.")
        sys.exit(1)

    ext = os.path.splitext(filename)[1].lower()
    if ext == '.xlsx':
        return read_excel_file(filename)
    else:
        return read_csv_file(filename)

# ---------- Загрузка конкретных сущностей ----------
def load_users() -> List[User]:
    data = load_data_file_interactive("inputDataUsers", "пользователей")
    if not data:
        messagebox.showerror("Ошибка", "Файл с пользователями пуст или не выбран.")
        sys.exit(1)

    users = []
    errors = []
    for idx, row in enumerate(data, start=2):
        try:
            # Ищем поля в любом регистре
            user_id = None
            for key in ['userid', 'user_id', 'id']:
                if key in row and row[key] and str(row[key]).strip():
                    user_id = int(str(row[key]).strip())
                    break
            if user_id is None:
                raise ValueError("Не найден идентификатор пользователя (userID)")

            fio = row.get('fio', '').strip()
            phone = str(row.get('phone', '')).strip()
            login = row.get('login', '').strip()
            password = row.get('password', '').strip()
            role_str = row.get('type', row.get('role', '')).strip()

            if not fio or not login or not password or not role_str:
                raise ValueError("Отсутствуют обязательные поля (fio, login, password, type)")

            users.append(User(
                user_id=user_id,
                fio=fio,
                phone=phone,
                login=login,
                password=password,
                role=UserRole(role_str)
            ))
        except Exception as e:
            errors.append(f"Строка {idx}: {row} -> {e}")

    if errors:
        print("\n".join(errors))
        messagebox.showerror("Ошибка",
            f"Не удалось прочитать {len(errors)} записей пользователей.\n"
            "Подробности в консоли.")
    if not users:
        messagebox.showerror("Ошибка", "Не удалось загрузить ни одного пользователя. Проверьте файл.")
        sys.exit(1)
    return users

def load_requests() -> List[Request]:
    data = load_data_file_interactive("inputDataRequests", "заявок")
    if not data:
        messagebox.showerror("Ошибка", "Файл с заявками пуст или не выбран.")
        sys.exit(1)

    requests = []
    errors = []
    for idx, row in enumerate(data, start=2):
        try:
            row_low = {k.lower(): v for k, v in row.items()}

            request_id = int(row_low.get('requestid', row_low.get('id', 0)))
            start_date_str = row_low.get('startdate', '')
            car_type = row_low.get('cartype', '').strip()
            car_model = row_low.get('carmodel', '').strip()
            problem_desc = row_low.get('problemdescryption', row_low.get('description', '')).strip()
            status_str = row_low.get('requeststatus', row_low.get('status', '')).strip()
            completion_date_str = row_low.get('completiondate', '')
            repair_parts = row_low.get('repairparts', '') or None
            master_id_str = row_low.get('masterid', '')
            client_id = int(row_low.get('clientid', 0))

            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            completion_date = None
            if completion_date_str and completion_date_str.lower() not in ('null', ''):
                completion_date = datetime.strptime(completion_date_str, '%Y-%m-%d').date()
            master_id = None
            if master_id_str and master_id_str.lower() not in ('null', ''):
                master_id = int(master_id_str)

            requests.append(Request(
                request_id=request_id,
                start_date=start_date,
                car_type=car_type,
                car_model=car_model,
                problem_desc=problem_desc,
                status=RequestStatus(status_str),
                completion_date=completion_date,
                repair_parts=repair_parts,
                master_id=master_id,
                client_id=client_id
            ))
        except Exception as e:
            errors.append(f"Строка {idx}: {row} -> {e}")

    if errors:
        print("\n".join(errors))
        messagebox.showerror("Ошибка",
            f"Не удалось прочитать {len(errors)} записей заявок.\n"
            "Подробности в консоли.")
    if not requests:
        messagebox.showerror("Ошибка", "Не удалось загрузить ни одной заявки. Проверьте файл.")
        sys.exit(1)
    return requests

def load_comments() -> List[Comment]:
    data = load_data_file_interactive("inputDataComments", "комментариев")
    if not data:
        messagebox.showerror("Ошибка", "Файл с комментариями пуст или не выбран.")
        sys.exit(1)

    comments = []
    errors = []
    for idx, row in enumerate(data, start=2):
        try:
            row_low = {k.lower(): v for k, v in row.items()}
            comment_id = int(row_low.get('commentid', 0))
            message = row_low.get('message', '').strip()
            master_id = int(row_low.get('masterid', 0))
            request_id = int(row_low.get('requestid', 0))
            comments.append(Comment(comment_id, message, master_id, request_id))
        except Exception as e:
            errors.append(f"Строка {idx}: {row} -> {e}")

    if errors:
        print("\n".join(errors))
    return comments

# ---------- Окно авторизации ----------
class LoginWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Авторизация - Учёт заявок на ремонт")
        self.geometry("300x200")
        self.resizable(False, False)

        # Загружаем пользователей (без этого не запустится)
        try:
            self.users = load_users()
        except Exception as e:
            messagebox.showerror("Критическая ошибка", str(e))
            sys.exit(1)

        self.current_user = None

        # Виджеты
        tk.Label(self, text="Логин:").pack(pady=(20,5))
        self.entry_login = tk.Entry(self)
        self.entry_login.pack(pady=5)
        self.entry_login.focus()

        tk.Label(self, text="Пароль:").pack(pady=5)
        self.entry_password = tk.Entry(self, show="*")
        self.entry_password.pack(pady=5)

        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=20)
        tk.Button(btn_frame, text="Войти", command=self.check_login, width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Выход", command=self.quit, width=10).pack(side=tk.LEFT, padx=5)

        # Привязка Enter
        self.entry_password.bind('<Return>', lambda e: self.check_login())

    def check_login(self):
        login = self.entry_login.get().strip()
        password = self.entry_password.get().strip()

        if not login or not password:
            messagebox.showwarning("Предупреждение", "Введите логин и пароль.")
            return

        # Ищем пользователя
        found_user = None
        for u in self.users:
            if u.login == login and u.password == password:
                found_user = u
                break

        if found_user:
            self.current_user = found_user
            self.destroy()  # закрываем окно авторизации
        else:
            messagebox.showerror("Ошибка", "Неверный логин или пароль.")
            self.entry_password.delete(0, tk.END)
            self.entry_login.focus()

# ---------- Главное окно приложения ----------
class MainWindow(tk.Tk):
    def __init__(self, current_user: User):
        super().__init__()
        self.current_user = current_user
        self.title(f"Учёт заявок на ремонт автомобилей - {current_user.fio} ({current_user.role.value})")
        self.geometry("1000x500")

        # Загрузка остальных данных (заявки, комментарии)
        try:
            self.users = load_users()  # уже загружены, но для полноты
            self.requests = load_requests()
            self.comments = load_comments()
        except Exception as e:
            messagebox.showerror("Критическая ошибка", str(e))
            sys.exit(1)

        self.next_request_id = max((r.request_id for r in self.requests), default=0) + 1
        self.next_comment_id = max((c.comment_id for c in self.comments), default=0) + 1

        self._create_widgets()
        self._refresh_table()

    def _create_widgets(self):
        # Таблица заявок
        columns = ("id", "client", "car", "status", "start", "end", "master")
        self.tree = ttk.Treeview(self, columns=columns, show="headings")
        self.tree.heading("id", text="№")
        self.tree.heading("client", text="Клиент")
        self.tree.heading("car", text="Автомобиль")
        self.tree.heading("status", text="Статус")
        self.tree.heading("start", text="Дата приёма")
        self.tree.heading("end", text="Дата завершения")
        self.tree.heading("master", text="Мастер")
        self.tree.column("id", width=50)
        self.tree.column("client", width=150)
        self.tree.column("car", width=150)
        self.tree.column("status", width=150)
        self.tree.column("start", width=100)
        self.tree.column("end", width=100)
        self.tree.column("master", width=150)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Кнопки
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(btn_frame, text="Добавить", command=self.add_request).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Редактировать", command=self.edit_request).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Удалить", command=self.delete_request).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Комментарии", command=self.show_comments).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Статистика", command=self.show_statistics).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Выход", command=self.logout).pack(side=tk.RIGHT, padx=2)

        # Строка статуса
        self.status_bar = tk.Label(self, text=f"Вы вошли как: {self.current_user.fio} ({self.current_user.role.value})", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def logout(self):
        self.destroy()
        # Возвращаемся к окну авторизации
        login = LoginWindow()
        login.mainloop()

    def _refresh_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for req in self.requests:
            client = next((u.fio for u in self.users if u.user_id == req.client_id), "Неизвестно")
            master = next((u.fio for u in self.users if u.user_id == req.master_id), "") if req.master_id else ""
            self.tree.insert("", tk.END, values=(
                req.request_id,
                client,
                f"{req.car_type} {req.car_model}",
                req.status.value,
                req.start_date.strftime("%d.%m.%Y"),
                req.completion_date.strftime("%d.%m.%Y") if req.completion_date else "",
                master
            ))

    def _get_selected_request(self) -> Optional[Request]:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Нет выбора", "Выберите заявку из списка.")
            return None
        item = self.tree.item(selected[0])
        req_id = int(item['values'][0])
        return next((r for r in self.requests if r.request_id == req_id), None)

    def add_request(self):
        RequestEditWindow(self, None)

    def edit_request(self):
        req = self._get_selected_request()
        if req:
            RequestEditWindow(self, req)

    def delete_request(self):
        req = self._get_selected_request()
        if not req:
            return
        confirm = messagebox.askyesno("Подтверждение", f"Удалить заявку №{req.request_id}?\nВсе связанные комментарии также будут удалены.")
        if confirm:
            self.requests.remove(req)
            self.comments = [c for c in self.comments if c.request_id != req.request_id]
            self._refresh_table()
            messagebox.showinfo("Успешно", "Заявка удалена.")

    def show_comments(self):
        req = self._get_selected_request()
        if req:
            CommentsWindow(self, req)

    def show_statistics(self):
        total = len(self.requests)
        status_counts = {s: 0 for s in RequestStatus}
        for req in self.requests:
            status_counts[req.status] += 1
        avg_time = self.calc_avg_repair_time()
        msg = f"Всего заявок: {total}\n"
        for status, count in status_counts.items():
            msg += f"{status.value}: {count}\n"
        if avg_time > 0:
            msg += f"Среднее время ремонта (дней): {avg_time:.1f}"
        else:
            msg += "Нет завершённых заявок для расчёта среднего времени."
        messagebox.showinfo("Статистика", msg)

    def calc_avg_repair_time(self) -> float:
        total_days = 0
        count = 0
        for req in self.requests:
            if req.status == RequestStatus.READY and req.completion_date:
                delta = (req.completion_date - req.start_date).days
                total_days += delta
                count += 1
        return total_days / count if count > 0 else 0.0

# ---------- Окно редактирования заявки ----------
class RequestEditWindow(tk.Toplevel):
    def __init__(self, parent: MainWindow, request: Optional[Request]):
        super().__init__(parent)
        self.parent = parent
        self.request = request
        self.title("Редактирование заявки" if request else "Новая заявка")
        self.geometry("500x400")
        self.resizable(False, False)

        self.customers = [u for u in parent.users if u.role == UserRole.CUSTOMER]
        self.mechanics = [u for u in parent.users if u.role == UserRole.MECHANIC]

        self.client_var = tk.StringVar()
        self.car_type_var = tk.StringVar()
        self.car_model_var = tk.StringVar()
        self.desc_var = tk.StringVar()
        self.status_var = tk.StringVar()
        self.master_var = tk.StringVar()
        self.start_date_var = tk.StringVar()
        self.compl_date_var = tk.StringVar()

        if request:
            self.client_var.set(str(request.client_id))
            self.car_type_var.set(request.car_type)
            self.car_model_var.set(request.car_model)
            self.desc_var.set(request.problem_desc)
            self.status_var.set(request.status.value)
            self.master_var.set(str(request.master_id) if request.master_id else "")
            self.start_date_var.set(request.start_date.strftime("%d.%m.%Y"))
            self.compl_date_var.set(request.completion_date.strftime("%d.%m.%Y") if request.completion_date else "")
        else:
            self.start_date_var.set(date.today().strftime("%d.%m.%Y"))

        self._create_widgets()

    def _create_widgets(self):
        row = 0
        tk.Label(self, text="Клиент:*").grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
        client_combo = ttk.Combobox(self, textvariable=self.client_var, values=[f"{c.user_id}: {c.fio}" for c in self.customers], state="readonly")
        client_combo.grid(row=row, column=1, padx=5, pady=2, sticky=tk.W)
        row += 1

        tk.Label(self, text="Тип авто:*").grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Entry(self, textvariable=self.car_type_var).grid(row=row, column=1, padx=5, pady=2, sticky=tk.W)
        row += 1

        tk.Label(self, text="Модель:*").grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Entry(self, textvariable=self.car_model_var).grid(row=row, column=1, padx=5, pady=2, sticky=tk.W)
        row += 1

        tk.Label(self, text="Описание:*").grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Entry(self, textvariable=self.desc_var, width=40).grid(row=row, column=1, padx=5, pady=2, sticky=tk.W)
        row += 1

        tk.Label(self, text="Статус:*").grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
        status_combo = ttk.Combobox(self, textvariable=self.status_var, values=[s.value for s in RequestStatus], state="readonly")
        status_combo.grid(row=row, column=1, padx=5, pady=2, sticky=tk.W)
        row += 1

        tk.Label(self, text="Мастер:").grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
        master_combo = ttk.Combobox(self, textvariable=self.master_var, values=[""] + [f"{m.user_id}: {m.fio}" for m in self.mechanics], state="readonly")
        master_combo.grid(row=row, column=1, padx=5, pady=2, sticky=tk.W)
        row += 1

        tk.Label(self, text="Дата приёма (ДД.ММ.ГГГГ):*").grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Entry(self, textvariable=self.start_date_var).grid(row=row, column=1, padx=5, pady=2, sticky=tk.W)
        row += 1

        tk.Label(self, text="Дата завершения:").grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
        tk.Entry(self, textvariable=self.compl_date_var).grid(row=row, column=1, padx=5, pady=2, sticky=tk.W)
        row += 1

        btn_frame = tk.Frame(self)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=10)
        tk.Button(btn_frame, text="Сохранить", command=self.save).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Отмена", command=self.destroy).pack(side=tk.LEFT, padx=5)

    def save(self):
        if not self.client_var.get() or not self.car_type_var.get() or not self.car_model_var.get() or not self.desc_var.get() or not self.status_var.get() or not self.start_date_var.get():
            messagebox.showerror("Ошибка", "Поля, отмеченные *, обязательны для заполнения.")
            return

        try:
            start = datetime.strptime(self.start_date_var.get(), "%d.%m.%Y").date()
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты приёма. Используйте ДД.ММ.ГГГГ.")
            return

        compl = None
        if self.compl_date_var.get().strip():
            try:
                compl = datetime.strptime(self.compl_date_var.get(), "%d.%m.%Y").date()
                if compl < start:
                    messagebox.showerror("Ошибка", "Дата завершения не может быть раньше даты приёма.")
                    return
            except ValueError:
                messagebox.showerror("Ошибка", "Неверный формат даты завершения.")
                return

        client_id = int(self.client_var.get().split(":")[0])
        master_id = None
        if self.master_var.get().strip():
            master_id = int(self.master_var.get().split(":")[0])

        status = RequestStatus(self.status_var.get())

        if self.request:
            self.request.client_id = client_id
            self.request.car_type = self.car_type_var.get()
            self.request.car_model = self.car_model_var.get()
            self.request.problem_desc = self.desc_var.get()
            self.request.status = status
            self.request.master_id = master_id
            self.request.start_date = start
            self.request.completion_date = compl
        else:
            new_id = self.parent.next_request_id
            self.parent.next_request_id += 1
            new_req = Request(
                request_id=new_id,
                start_date=start,
                car_type=self.car_type_var.get(),
                car_model=self.car_model_var.get(),
                problem_desc=self.desc_var.get(),
                status=status,
                completion_date=compl,
                repair_parts=None,
                master_id=master_id,
                client_id=client_id
            )
            self.parent.requests.append(new_req)

        self.parent._refresh_table()
        messagebox.showinfo("Успешно", "Заявка сохранена.")
        self.destroy()

# ---------- Окно комментариев ----------
class CommentsWindow(tk.Toplevel):
    def __init__(self, parent: MainWindow, request: Request):
        super().__init__(parent)
        self.parent = parent
        self.request = request
        self.title(f"Комментарии к заявке №{request.request_id}")
        self.geometry("500x400")

        self.listbox = tk.Listbox(self)
        self.listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        entry_frame = tk.Frame(self)
        entry_frame.pack(fill=tk.X, padx=5, pady=5)
        self.comment_entry = tk.Entry(entry_frame, width=40)
        self.comment_entry.pack(side=tk.LEFT, padx=2)
        tk.Button(entry_frame, text="Добавить", command=self.add_comment).pack(side=tk.LEFT, padx=2)
        tk.Button(entry_frame, text="Назад", command=self.destroy).pack(side=tk.RIGHT, padx=2)

        self._refresh_comments()

    def _refresh_comments(self):
        self.listbox.delete(0, tk.END)
        for c in self.parent.comments:
            if c.request_id == self.request.request_id:
                master = next((u.fio for u in self.parent.users if u.user_id == c.master_id), "Неизвестно")
                self.listbox.insert(tk.END, f"{master}: {c.message}")

    def add_comment(self):
        text = self.comment_entry.get().strip()
        if not text:
            messagebox.showwarning("Пустой комментарий", "Введите текст комментария.")
            return
        mechanic = next((u for u in self.parent.users if u.role == UserRole.MECHANIC), None)
        if not mechanic:
            messagebox.showerror("Ошибка", "Нет автомехаников в системе.")
            return
        new_id = self.parent.next_comment_id
        self.parent.next_comment_id += 1
        comment = Comment(
            comment_id=new_id,
            message=text,
            master_id=mechanic.user_id,
            request_id=self.request.request_id
        )
        self.parent.comments.append(comment)
        self.comment_entry.delete(0, tk.END)
        self._refresh_comments()

# ---------- Запуск ----------
if __name__ == "__main__":
    login_win = LoginWindow()
    login_win.mainloop()
    if login_win.current_user:
        app = MainWindow(login_win.current_user)
        app.mainloop()